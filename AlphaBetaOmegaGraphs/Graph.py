
#? This Program is Designed to Read the dall2020.dat input file and look for the alphas, omegas and betas.
#? Upon reading them they are then graphed.
#* Created by: Evan Petrimoulx May 30 2023

import matplotlib.pyplot as plt
import numpy as np
import os
import openpyxl
from pathlib import Path
from openpyxl.chart import Reference, LineChart


def getAlphasBetasOmegas():
    # open files for reading and writing
    dall2020 = open('dall2020.dat', 'r')
    storedNumbers = open('StoredNumbers.txt', 'w+')

    #grab content of dall2020.dat file
    content = dall2020.readlines()
    
    #grab alphas betas and omegas from dall2020 and put them into the StoredNumbers.txt file
    for i in range(0, len(content)):
        if('OK' in content[i] or 'NDG' in content[i]):
            AlphasBetas = content[i+1].split(',')
            storedNumbers.write(AlphasBetas[0][1:] + ', ' + AlphasBetas[1] + ', ' + AlphasBetas[2] + ', ' + AlphasBetas[3] + ', ' + AlphasBetas[4] + ', ' + AlphasBetas[5] + ', ')
            Omegas = content[i+2].split(',')
            storedNumbers.write(Omegas[0] + ', ' + Omegas[1] + ', ' + Omegas[2] + '\n')
        
    # close files
    storedNumbers.close()
    dall2020.close()
    return 0
    
def makeArray():
    #open storedNumbers to grab the data and graph it
    storedNumbers = open('StoredNumbers.txt', 'r')
    
    #grab content of storedNumbers.txt file
    content = storedNumbers.readlines()
    array = []
    
    #putting the information stored in content (AKA the info from StoredNumbers.txt) into the array - creating an array of lists
    for i in range(0, len(content)):
        array.append(content[i].split(','))
        
    #removing ,'s from list as well as newline characters so we can convert it to type float
    for i in range(0, len(content)):
        for j in range(0, len(array[i])):
            array[i][j] = float(array[i][j].replace(" ", "").replace("\n", ""))
            
            
    #take the transpose of the array of alphas betas and omegas to make it easier to handle for graphing (All alpha_1's beta_1's etc in the same row)
    array = np.transpose(array)

    # close file
    storedNumbers.close()
    
    return array
    
def makePythonGraph(array):
    
    #Read values from the stored array and assign them to their proper variables
    alpha_1 = array[0][:]
    beta_1 = array[1][:]
    alpha_2 = array[2][:]
    beta_2 = array[3][:]
    alpha_3 = array[4][:]
    beta_3 = array[5][:]
    Omega_1 = array[6][:]
    Omega_2 = array[7][:]
    Omega_3 = array[8][:]
    
    plt.plot(Omega_1, alpha_1, color = 'blue', label = 'Alpha 1', linestyle = 'solid', marker = 'o', linewidth = 2, markerfacecolor = 'blue', markersize = 5)
    plt.plot(Omega_1, alpha_2, color = 'blue', label = 'Alpha 2', linestyle = 'solid', marker = 'o', linewidth = 2, markerfacecolor = 'blue', markersize = 5)
    plt.plot(Omega_1, alpha_3, color = 'blue', label = 'Alpha 3', linestyle = 'solid', marker = 'o', linewidth = 2, markerfacecolor = 'blue', markersize = 5)
    plt.plot(Omega_1, beta_1, color = 'orange', label = 'Beta 1', linestyle = 'solid', marker = 'o', linewidth = 2, markerfacecolor = 'orange', markersize = 5)
    plt.plot(Omega_1, beta_2, color = 'orange', label = 'Beta 2', linestyle = 'solid', marker = 'o', linewidth = 2, markerfacecolor = 'orange', markersize = 5)
    plt.plot(Omega_1, beta_3, color = 'orange', label = 'Beta 3', linestyle = 'solid', marker = 'o', linewidth = 2, markerfacecolor = 'orange', markersize = 5)
    
    #Labelling the Graphs and adding titles
    plt.xticks(np.arange(min(Omega_1), max(Omega_1)+1, 1.0))
    plt.xlabel('Omega 1')
    plt.ylabel('Alphas and Betas')
    plt.title("Plot of Alphas and Betas vs Omegas")
    plt.legend()
    plt.show()
    return 0

def makeExcelGraph(array):
    Titles = [
        "Alpha 1",
        "Alpha 2",
        "Alpha 3",
        "Beta 1",
        "Beta 2",
        "Beta 3",
        "Omega 1",
        "Omega 2",
        "Omega 3",
    ]
    #Switch the columns and rows back to match the generated text file from earlier. This makes it easier for Excel to read the data
    array = np.transpose(array)
    
    #Path to ExcelSheet
    path = str(Path.cwd())
    path = path + '/ExcelGraphs.xlsx'
    
    #Check if the ExcelSheet exists in your directory, if it does, load it, if it does not, create one
    if(os.path.isfile('ExcelGraphs.xlsx')):
        ExcelGraphBook = openpyxl.load_workbook(path)
    else:
        ExcelGraphBook = openpyxl.Workbook('ExcelGraphs.xlsx')
        ExcelGraphBook.save(filename = 'ExcelGraphs.xlsx')

    #Open a new sheet
    NumberOfSheets = len(ExcelGraphBook.sheetnames) + 1
    for i in range(0, NumberOfSheets + 1):
        SheetName = f"Graph-{i+1}"
        try:
            ExcelGraphBook[SheetName]
        except:
            break
    
    Sheet = ExcelGraphBook.create_sheet(SheetName)
    
    #Put Array Data into Excel Sheet
    for i in range(len(array)):
        for j in range(len(array[0])):
            Numbers = Sheet.cell(row = i+2, column = j+1)
            Numbers.value = array[i][j] 

    for i in range(Sheet.max_column):
        Sheet.cell(row = 1, column = i+1).value = Titles[i]

    #Variables for range of values
    Alpha_1_Start = 'A1'
    Alpha_1_End = f'A{Sheet.max_row}'
    Alpha_2_Start = 'B1'
    Alpha_2_End = f'B{Sheet.max_row}'
    Alpha_3_Start = 'C1'
    Alpha_3_End = f'C{Sheet.max_row}'
    Beta_1_Start = 'D1'
    Beta_1_End = f'D{Sheet.max_row}'
    Beta_2_Start = 'E1'
    Beta_2_End = f'E{Sheet.max_row}'
    Beta_3_Start = 'F1'
    Beta_3_End = f'F{Sheet.max_row}'

    Omega_1_Start = 'G2'
    Omega_1_End = f'G{Sheet.max_row}'
    Omega_2_Start = 'H2'
    Omega_2_End = f'H{Sheet.max_row}'
    Omega_3_Start = 'I2'
    Omega_3_End = f'I{Sheet.max_row}'
    
    #Initialize Chart
    Chart = LineChart()


    #Initialize references for data based on the ranges of values from above
    Alpha_1_Values = Reference(Sheet, range_string=f"{SheetName}!{Alpha_1_Start}:{Alpha_1_End}")
    Alpha_2_Values = Reference(Sheet, range_string=f"{SheetName}!{Alpha_2_Start}:{Alpha_2_End}")
    Alpha_3_Values = Reference(Sheet, range_string=f"{SheetName}!{Alpha_3_Start}:{Alpha_3_End}")
    Beta_1_Values = Reference(Sheet, range_string=f"{SheetName}!{Beta_1_Start}:{Beta_1_End}")
    Beta_2_Values = Reference(Sheet, range_string=f"{SheetName}!{Beta_2_Start}:{Beta_2_End}")
    Beta_3_Values = Reference(Sheet, range_string=f"{SheetName}!{Beta_3_Start}:{Beta_3_End}")

    Omega_1_Values = Reference(Sheet, range_string=f"{SheetName}!{Omega_1_Start}:{Omega_1_End}")
    Omega_2_Values = Reference(Sheet, range_string=f"{SheetName}!{Omega_2_Start}:{Omega_2_End}")
    Omega_3_Values = Reference(Sheet, range_string=f"{SheetName}!{Omega_3_Start}:{Omega_3_End}")


    #Add data to the chart and set the x axis to Omega 1 values
    Chart.add_data(Alpha_1_Values, titles_from_data = True)
    Chart.add_data(Alpha_2_Values, titles_from_data = True)
    Chart.add_data(Alpha_3_Values, titles_from_data = True)
    
    Chart.add_data(Beta_1_Values, titles_from_data = True)
    Chart.add_data(Beta_2_Values, titles_from_data = True)
    Chart.add_data(Beta_3_Values, titles_from_data = True)

    Chart.set_categories(Omega_1_Values)

    #Cosmetics
    Chart.title = 'Plot of Alphas and Betas vs Omegas'
    Chart.x_axis.title = 'Omega'
    Chart.y_axis.title = 'Alpha and Beta Values'
    Chart.legend.position = 'b'
    Chart.height = 12
    Chart.width = 20
    Chart.style = 5
    
    #^ FIXME::       Chart.legend.Legend(legendEntry = (['Alpha 1', 'Alpha 2', 'Alpha 3', 'Beta 1', 'Beta 2', 'Beta 3']))
    
    #Make Graph
    Sheet.add_chart(Chart, 'K1')
    
    #Save Worksheet
    ExcelGraphBook.save('ExcelGraphs.xlsx')
    
    return 0

getAlphasBetasOmegas()
makeExcelGraph(makeArray())
makePythonGraph(makeArray())

#Delete the .txt when you are done with it, you can comment this out if you want to view its contents, you can also view the same contents by looking at the excel chart
os.remove('StoredNumbers.txt')

#TODO - Check to see if the bumps and unexpected wiggles are real. (Restart the program )
#TODO - Make a new graph of the extrapolated value and the difference for any omega (Eomega - Eextrapolated) -- should be a logarithmic plot